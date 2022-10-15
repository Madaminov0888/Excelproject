import openpyxl
from openpyxl import load_workbook
import datetime

workbook = load_workbook('desktop/tgt/rezhim.xlsx')
sheet = workbook.active
for i in range(4, sheet.max_row+1):  #sheet.max_row+1
    h = []
    for j in range(2, 29, 2):
        u = []
        wrk_time = sheet.cell(i, j).value.split('-')
        dam_time = sheet.cell(i, j+1).value.split('-')
        if len(wrk_time) < 2:
            u.extend([0, 0, 0])
            h.append(u)
        else:
            wrk_minut = datetime.timedelta(hours=int(wrk_time[1].split(':')[0]), minutes=int(wrk_time[1].split(':')[1])) - datetime.timedelta(hours=int(wrk_time[0].split(':')[0]), minutes=int(wrk_time[0].split(':')[1]))
            min1 = divmod(wrk_minut.seconds, 60)[0]
            if len(dam_time) < 2:
                min2 = 0
            else:
                dam_minut = datetime.timedelta(hours=int(dam_time[1].split(':')[0]), minutes=int(dam_time[1].split(':')[1])) - datetime.timedelta(hours=int(dam_time[0].split(':')[0]), minutes=int(dam_time[0].split(':')[1]))
                min2 = divmod(dam_minut.seconds, 60)[0]
            u.extend([min1, min2, min1-min2])
            h.append(u)
    
    for k in range(30, 72, 3):
        index = k//3-10
        sheet.cell(i, k).value = h[index][0]
        sheet.cell(i, k+1).value = h[index][1]
        sheet.cell(i, k+2).value = h[index][2]
    
    for last in range(72, 79):
        index1 = last - 72
        sheet.cell(i, last).value = h[index1][2] - h[index1+7][2]


workbook.save('desktop/tgt/rezhim1.xlsx')
        
        
